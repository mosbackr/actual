"""Generate xlsx of startup/VC publications by city."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

data = [
    # US Tier 1
    ("San Francisco", "CA", "US", "TechCrunch", "https://techcrunch.com/"),
    ("San Francisco", "CA", "US", "Built In San Francisco", "https://www.builtinsf.com/"),
    ("San Francisco", "CA", "US", "SF Business Times", "https://www.bizjournals.com/sanfrancisco/"),
    ("New York", "NY", "US", "AlleyWatch", "https://www.alleywatch.com/"),
    ("New York", "NY", "US", "Built In NYC", "https://www.builtinnyc.com/"),
    ("New York", "NY", "US", "Crain's New York Business", "https://www.crainsnewyork.com/"),
    ("Boston", "MA", "US", "BostInno", "https://www.bizjournals.com/boston/inno"),
    ("Boston", "MA", "US", "Built In Boston", "https://www.builtinboston.com/"),
    ("Boston", "MA", "US", "Boston Business Journal", "https://www.bizjournals.com/boston/"),
    ("Los Angeles", "CA", "US", "dot.LA", "https://dot.la/"),
    ("Los Angeles", "CA", "US", "LA TechWatch", "https://www.latechwatch.com/"),
    ("Los Angeles", "CA", "US", "Built In LA", "https://www.builtinla.com/"),
    ("Seattle", "WA", "US", "GeekWire", "https://www.geekwire.com/"),
    ("Seattle", "WA", "US", "Built In Seattle", "https://www.builtinseattle.com/"),
    ("Seattle", "WA", "US", "Puget Sound Business Journal", "https://www.bizjournals.com/seattle/"),
    ("Austin", "TX", "US", "Silicon Hills News", "https://www.siliconhillsnews.com/"),
    ("Austin", "TX", "US", "Built In Austin", "https://www.builtinaustin.com/"),
    ("Austin", "TX", "US", "Austin Business Journal", "https://www.bizjournals.com/austin/"),
    ("Chicago", "IL", "US", "Built In Chicago", "https://www.builtinchicago.org/"),
    ("Chicago", "IL", "US", "Chicago Inno", "https://www.bizjournals.com/chicago/inno"),
    ("Chicago", "IL", "US", "Crain's Chicago Business", "https://www.chicagobusiness.com/"),
    ("Miami", "FL", "US", "Refresh Miami", "https://refreshmiami.com/"),
    ("Miami", "FL", "US", "Miami Inno", "https://www.bizjournals.com/southflorida/inno"),
    ("Miami", "FL", "US", "South Florida Business Journal", "https://www.bizjournals.com/southflorida/"),
    ("Denver", "CO", "US", "Built In Colorado", "https://www.builtincolorado.com/"),
    ("Denver", "CO", "US", "Denver Business Journal", "https://www.bizjournals.com/denver/"),
    ("Denver", "CO", "US", "Colorado Sun", "https://coloradosun.com/"),
    ("Washington", "DC", "US", "Technical.ly DC", "https://technical.ly/dc/"),
    ("Washington", "DC", "US", "DC Inno", "https://www.bizjournals.com/washington/inno"),
    ("Washington", "DC", "US", "Washington Business Journal", "https://www.bizjournals.com/washington/"),
    ("San Jose", "CA", "US", "Silicon Valley Business Journal", "https://www.bizjournals.com/sanjose/"),
    ("San Jose", "CA", "US", "The Information", "https://www.theinformation.com/"),
    ("San Jose", "CA", "US", "San Jose Spotlight", "https://sanjosespotlight.com/"),
    # US Tier 2
    ("San Diego", "CA", "US", "San Diego Business Journal", "https://www.sdbj.com/"),
    ("San Diego", "CA", "US", "Startup San Diego", "https://startupsd.org/"),
    ("San Diego", "CA", "US", "Times of San Diego", "https://timesofsandiego.com/"),
    ("Atlanta", "GA", "US", "Hypepotamus", "https://hypepotamus.com/"),
    ("Atlanta", "GA", "US", "Atlanta Inno", "https://www.bizjournals.com/atlanta/inno"),
    ("Atlanta", "GA", "US", "Atlanta Business Chronicle", "https://www.bizjournals.com/atlanta/"),
    ("Dallas", "TX", "US", "Dallas Innovates", "https://dallasinnovates.com/"),
    ("Dallas", "TX", "US", "North Texas Inno", "https://www.bizjournals.com/dallas/inno"),
    ("Dallas", "TX", "US", "Dallas Business Journal", "https://www.bizjournals.com/dallas/"),
    ("Houston", "TX", "US", "InnovationMap", "https://houston.innovationmap.com/"),
    ("Houston", "TX", "US", "Houston Business Journal", "https://www.bizjournals.com/houston/"),
    ("Houston", "TX", "US", "Houston Chronicle Tech", "https://www.houstonchronicle.com/business/technology/"),
    ("Philadelphia", "PA", "US", "Technical.ly Philly", "https://technical.ly/philly/"),
    ("Philadelphia", "PA", "US", "Philly Inno", "https://www.bizjournals.com/philadelphia/inno"),
    ("Philadelphia", "PA", "US", "Philadelphia Business Journal", "https://www.bizjournals.com/philadelphia/"),
    ("Minneapolis", "MN", "US", "Twin Cities Business", "https://tcbmag.com/"),
    ("Minneapolis", "MN", "US", "Minneapolis/St. Paul Business Journal", "https://www.bizjournals.com/twincities/"),
    ("Minneapolis", "MN", "US", "MnTech", "https://mntech.org/"),
    ("Detroit", "MI", "US", "MITechNews", "https://mitechnews.com/"),
    ("Detroit", "MI", "US", "Crain's Detroit Business", "https://www.crainsdetroit.com/"),
    ("Detroit", "MI", "US", "TechTown Detroit", "https://techtowndetroit.org/"),
    ("Pittsburgh", "PA", "US", "Technical.ly Pittsburgh", "https://technical.ly/pittsburgh/"),
    ("Pittsburgh", "PA", "US", "Pittsburgh Tech Council", "https://www.pghtech.org/"),
    ("Nashville", "TN", "US", "Nashville Post", "https://www.nashvillepost.com/"),
    ("Nashville", "TN", "US", "Nashville Business Journal", "https://www.bizjournals.com/nashville/"),
    ("Raleigh-Durham", "NC", "US", "WRAL TechWire", "https://www.wraltechwire.com/"),
    ("Raleigh-Durham", "NC", "US", "Triangle Business Journal", "https://www.bizjournals.com/triangle/"),
    ("Salt Lake City", "UT", "US", "Beehive Startups", "https://beehivestartups.com/"),
    ("Salt Lake City", "UT", "US", "TechBuzz News", "https://www.techbuzznews.com/"),
    ("Salt Lake City", "UT", "US", "Silicon Slopes", "https://www.siliconslopes.com/"),
    ("Portland", "OR", "US", "Silicon Florist", "https://siliconflorist.com/"),
    ("Portland", "OR", "US", "Portland Business Journal", "https://www.bizjournals.com/portland/"),
    ("Phoenix", "AZ", "US", "AZ Big Media", "https://azbigmedia.com/"),
    ("Phoenix", "AZ", "US", "Phoenix Business Journal", "https://www.bizjournals.com/phoenix/"),
    ("Columbus", "OH", "US", "Columbus Inno", "https://www.bizjournals.com/columbus/inno"),
    ("Columbus", "OH", "US", "Columbus Business First", "https://www.bizjournals.com/columbus/"),
    ("Indianapolis", "IN", "US", "TechPoint Index", "https://techpoint.org/"),
    ("Indianapolis", "IN", "US", "Indianapolis Business Journal", "https://www.ibj.com/"),
    ("Indianapolis", "IN", "US", "Inside Indiana Business", "https://www.insideindianabusiness.com/"),
    ("St. Louis", "MO", "US", "EQ STL", "https://www.eqstl.com/"),
    ("St. Louis", "MO", "US", "St. Louis Business Journal", "https://www.bizjournals.com/stlouis/"),
    ("Baltimore", "MD", "US", "Technical.ly Baltimore", "https://technical.ly/baltimore/"),
    ("Baltimore", "MD", "US", "Baltimore Business Journal", "https://www.bizjournals.com/baltimore/"),
    ("Tampa", "FL", "US", "St. Pete Catalyst", "https://stpetecatalyst.com/"),
    ("Tampa", "FL", "US", "Tampa Bay Inno", "https://www.bizjournals.com/tampabay/inno"),
    ("Tampa", "FL", "US", "Tampa Bay Business Journal", "https://www.bizjournals.com/tampabay/"),
    ("Charlotte", "NC", "US", "Charlotte Inno", "https://www.bizjournals.com/charlotte/inno"),
    ("Charlotte", "NC", "US", "Charlotte Business Journal", "https://www.bizjournals.com/charlotte/"),
    ("Las Vegas", "NV", "US", "Vegas Inc", "https://vegasinc.lasvegassun.com/"),
    ("Las Vegas", "NV", "US", "StartUpNV", "https://www.startupnv.com/"),
    ("Cincinnati", "OH", "US", "Cincinnati Inno", "https://www.bizjournals.com/cincinnati/inno"),
    ("Cincinnati", "OH", "US", "Cintrifuse", "https://www.cintrifuse.com/"),
    ("Cincinnati", "OH", "US", "Cincinnati Business Courier", "https://www.bizjournals.com/cincinnati/"),
    ("Kansas City", "MO", "US", "Startland News", "https://www.startlandnews.com/"),
    ("Kansas City", "MO", "US", "Silicon Prairie News", "https://siliconprairienews.com/"),
    ("Kansas City", "MO", "US", "Kansas City Business Journal", "https://www.bizjournals.com/kansascity/"),
    ("Birmingham", "AL", "US", "Alabama Inno", "https://www.bizjournals.com/birmingham/inno"),
    ("Birmingham", "AL", "US", "Birmingham Business Journal", "https://www.bizjournals.com/birmingham/"),
    ("Madison", "WI", "US", "In Business Madison", "https://www.ibmadison.com/"),
    ("Madison", "WI", "US", "Cap Times", "https://captimes.com/"),
    ("Omaha", "NE", "US", "Silicon Prairie News", "https://siliconprairienews.com/"),
    ("Omaha", "NE", "US", "Omaha World-Herald Business", "https://omaha.com/business/"),
    ("Sacramento", "CA", "US", "Comstock's Magazine", "https://www.comstocksmag.com/"),
    ("Sacramento", "CA", "US", "Sacramento Business Journal", "https://www.bizjournals.com/sacramento/"),
    ("Oakland", "CA", "US", "The Oaklandside", "https://oaklandside.org/"),
    ("Oakland", "CA", "US", "SF Business Times (East Bay)", "https://www.bizjournals.com/sanfrancisco/"),
    ("Orlando", "FL", "US", "Orlando Inno", "https://www.bizjournals.com/orlando/inno"),
    ("Orlando", "FL", "US", "Orlando Business Journal", "https://www.bizjournals.com/orlando/"),
    ("Cleveland", "OH", "US", "Cleveland Inno", "https://www.bizjournals.com/cleveland/inno"),
    ("Cleveland", "OH", "US", "Crain's Cleveland Business", "https://www.crainscleveland.com/"),
    ("Irvine", "CA", "US", "OCTANe", "https://octaneoc.org/"),
    ("Irvine", "CA", "US", "Orange County Business Journal", "https://www.ocbj.com/"),
    # US Tier 3
    ("San Antonio", "TX", "US", "San Antonio Business Journal", "https://www.bizjournals.com/sanantonio/"),
    ("San Antonio", "TX", "US", "Tech Bloc SA", "https://www.techbloc.org/"),
    ("Jacksonville", "FL", "US", "Jacksonville Business Journal", "https://www.bizjournals.com/jacksonville/"),
    ("Jacksonville", "FL", "US", "Jax Daily Record", "https://www.jaxdailyrecord.com/"),
    ("Fort Worth", "TX", "US", "Fort Worth Business Press", "https://www.fortworthbusiness.com/"),
    ("Fort Worth", "TX", "US", "Dallas-Fort Worth Inno", "https://www.bizjournals.com/dallas/inno"),
    ("Oklahoma City", "OK", "US", "The Journal Record", "https://journalrecord.com/"),
    ("Oklahoma City", "OK", "US", "Oklahoma Gazette", "https://www.okgazette.com/"),
    ("El Paso", "TX", "US", "El Paso Inc.", "https://www.elpasoinc.com/"),
    ("El Paso", "TX", "US", "El Paso Times Business", "https://www.elpasotimes.com/business/"),
    ("Memphis", "TN", "US", "Memphis Business Journal", "https://www.bizjournals.com/memphis/"),
    ("Memphis", "TN", "US", "EPIcenter Memphis", "https://www.epicentermemphis.org/"),
    ("Louisville", "KY", "US", "Louisville Business First", "https://www.bizjournals.com/louisville/"),
    ("Louisville", "KY", "US", "Insider Louisville", "https://insiderlouisville.com/"),
    ("Milwaukee", "WI", "US", "Milwaukee Business Journal", "https://www.bizjournals.com/milwaukee/"),
    ("Milwaukee", "WI", "US", "BizTimes Milwaukee", "https://biztimes.com/"),
    ("Albuquerque", "NM", "US", "Albuquerque Business First", "https://www.bizjournals.com/albuquerque/"),
    ("Albuquerque", "NM", "US", "Albuquerque Journal Business", "https://www.abqjournal.com/business/"),
    ("Tucson", "AZ", "US", "Inside Tucson Business", "https://www.insidetucsonbusiness.com/"),
    ("Tucson", "AZ", "US", "Tucson Sentinel", "https://www.tucsonsentinel.com/"),
    ("Fresno", "CA", "US", "The Business Journal (Central Valley)", "https://thebusinessjournal.com/"),
    ("Fresno", "CA", "US", "GV Wire", "https://gvwire.com/"),
    ("Colorado Springs", "CO", "US", "Colorado Springs Business Journal", "https://www.csbj.com/"),
    ("Virginia Beach", "VA", "US", "Inside Business Hampton Roads", "https://www.pilotonline.com/inside-business/"),
    ("Virginia Beach", "VA", "US", "Virginia Business", "https://www.virginiabusiness.com/"),
    ("Tulsa", "OK", "US", "Tulsa Innovation Labs", "https://www.tulsainnovationlabs.com/"),
    ("Tulsa", "OK", "US", "Tulsa World Business", "https://tulsaworld.com/business/"),
    ("New Orleans", "LA", "US", "Silicon Bayou News", "https://siliconbayounews.com/"),
    ("New Orleans", "LA", "US", "New Orleans CityBusiness", "https://neworleanscitybusiness.com/"),
    ("Wichita", "KS", "US", "Wichita Business Journal", "https://www.bizjournals.com/wichita/"),
    ("Honolulu", "HI", "US", "Pacific Business News", "https://www.bizjournals.com/pacific/"),
    ("Honolulu", "HI", "US", "Hawaii Business Magazine", "https://www.hawaiibusiness.com/"),
    ("Lexington", "KY", "US", "Lane Report", "https://www.lanereport.com/"),
    ("Lexington", "KY", "US", "Lexington Herald-Leader Business", "https://www.kentucky.com/news/business/"),
    ("Corpus Christi", "TX", "US", "Corpus Christi Caller-Times Business", "https://www.caller.com/business/"),
    ("Riverside", "CA", "US", "Inland Empire Business Journal", "https://iebj.com/"),
    ("Newark", "NJ", "US", "ROI-NJ", "https://www.roi-nj.com/"),
    ("Newark", "NJ", "US", "NJTechWeekly", "https://njtechweekly.com/"),
    ("Bakersfield", "CA", "US", "The Business Journal (Central Valley)", "https://thebusinessjournal.com/"),
    ("Stockton", "CA", "US", "Central Valley Business Journal", "https://www.centralvalleybusinessjournal.com/"),
    # International - North America
    ("Toronto", None, "Canada", "BetaKit", "https://betakit.com/"),
    ("Toronto", None, "Canada", "MaRS Discovery District", "https://www.marsdd.com/"),
    ("Vancouver", None, "Canada", "Techcouver", "https://techcouver.com/"),
    ("Vancouver", None, "Canada", "Vancouver Tech Journal", "https://vantechjournal.com/"),
    ("Montreal", None, "Canada", "BetaKit Montreal", "https://betakit.com/tag/montreal/"),
    # International - Europe
    ("London", None, "UK", "Sifted", "https://sifted.eu/"),
    ("London", None, "UK", "TechRound", "https://techround.co.uk/"),
    ("London", None, "UK", "Startups of London", "https://startupsoflondon.com/"),
    ("Berlin", None, "Germany", "Sifted Germany", "https://sifted.eu/countries/germany"),
    ("Berlin", None, "Germany", "Tech in Berlin", "https://www.techinberlin.com/"),
    ("Paris", None, "France", "Maddyness", "https://www.maddyness.com/"),
    ("Paris", None, "France", "French Tech Journal", "https://www.frenchtechjournal.com/"),
    ("Amsterdam", None, "Netherlands", "Silicon Canals", "https://siliconcanals.com/"),
    ("Amsterdam", None, "Netherlands", "Sifted Netherlands", "https://sifted.eu/countries/netherlands"),
    ("Stockholm", None, "Sweden", "Swedish Tech News", "https://www.swedishtechnews.com/"),
    ("Stockholm", None, "Sweden", "Nordic Startup News", "https://nordicstartupnews.com/"),
    # International - Asia-Pacific
    ("Singapore", None, "Singapore", "e27", "https://e27.co/"),
    ("Singapore", None, "Singapore", "Tech in Asia", "https://www.techinasia.com/"),
    ("Sydney", None, "Australia", "Startup Daily", "https://www.startupdaily.net/"),
    ("Sydney", None, "Australia", "StartupSmart", "https://www.smartcompany.com.au/startupsmart/"),
    ("Bangalore", None, "India", "Inc42", "https://inc42.com/"),
    ("Bangalore", None, "India", "YourStory", "https://yourstory.com/"),
    ("Tel Aviv", None, "Israel", "CTech (Calcalist)", "https://www.calcalistech.com/"),
    ("Tel Aviv", None, "Israel", "NoCamels", "https://nocamels.com/"),
    # International - Latin America
    ("Sao Paulo", None, "Brazil", "LABS News", "https://labsnews.com/en/"),
    ("Sao Paulo", None, "Brazil", "Contxto", "https://contxto.com/en/"),
    ("Mexico City", None, "Mexico", "Contxto", "https://contxto.com/en/"),
    ("Bogota", None, "Colombia", "Contxto", "https://contxto.com/en/"),
    ("Bogota", None, "Colombia", "LatAm Republic", "https://www.latamrepublic.com/"),
]

wb = Workbook()
ws = wb.active
ws.title = "Startup Publications by City"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="F28C28", end_color="F28C28", fill_type="solid")
tier_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
tier_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style="thin", color="E8E6E3"),
    right=Side(style="thin", color="E8E6E3"),
    top=Side(style="thin", color="E8E6E3"),
    bottom=Side(style="thin", color="E8E6E3"),
)

# Headers
headers = ["City", "State", "Country", "Publication", "URL"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Data
for row_idx, (city, state, country, pub, url) in enumerate(data, 2):
    ws.cell(row=row_idx, column=1, value=city).border = thin_border
    ws.cell(row=row_idx, column=2, value=state or "").border = thin_border
    ws.cell(row=row_idx, column=3, value=country).border = thin_border
    ws.cell(row=row_idx, column=4, value=pub).border = thin_border
    url_cell = ws.cell(row=row_idx, column=5, value=url)
    url_cell.hyperlink = url
    url_cell.font = Font(color="0563C1", underline="single")
    url_cell.border = thin_border

# Column widths
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 8
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 38
ws.column_dimensions["E"].width = 55

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:E{len(data) + 1}"

output_path = "/Users/leemosbacker/Downloads/startup_publications_by_city.xlsx"
wb.save(output_path)
print(f"Saved {len(data)} entries to {output_path}")
